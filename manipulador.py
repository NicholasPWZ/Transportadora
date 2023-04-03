#importando as classes criadas e bibliotecas para realizar a leitura dos dados fornecidos através do Excel
from Caminhao import PequenoPorte, MedioPorte, GrandePorte
from openpyxl import load_workbook, Workbook
import pandas as pd

#Realizando a leitura dos dados do excel e alocando em umas lista de segunda dimensão
df = pd.read_csv("Distancias.csv", sep=";")
df.to_excel("Distancias.xlsx", index=False)
wb = load_workbook("Distancias.xlsx")
ws = wb.active
excel = []
#Loop para ler uma linha completa do excel e alocar ela dentro de uma lista de segunda dimensão
for row in ws.iter_rows():
    linha_lista = []
    for cell in row:
        linha_lista.append(cell.value)
    excel.append(linha_lista)

#Criando os caminhões, através das classes com as informações necessárias e distintas de cada modelo
pp, mp, gp = PequenoPorte(), MedioPorte(), GrandePorte()

#Criando o dicionário, que armazenará os cadastros relizados e poderá ser exibido ao usuário, caso solicitado
dicionario_carregamentos = {}

#Início do programa

    #Primeira funcionalidade, relizar a consulta de um trajeto com um modelo de caminhão.
def consulta():
        #Solicitando o nome da cidade e transformando em UpperCase para consultar através de match na lista com os dados do Excel
        cidadePartida = input('Informe a cidade de partida: ').upper()
        cidadeDestino = input('Informe a cidade de destino: ').upper()
        
       
        #Capturando o index das cidades dentro da lista com os dados, e validando se a cidade está na lista dos dados
        try:
            indexcidadeP = excel[0].index(cidadePartida)
            indexcidadeD = excel[0].index(cidadeDestino)
        except:
            Exception('Você informou uma cidade incorreta.')
        #Capturando a distância entre as cidades informadas    
        distanciaCidades = excel[indexcidadeP + 1][indexcidadeD]
         
         
        caminhao = input(
            'Qual caminhão será utilizado?\n'
            '1- Pequeno Porte\n'
            '2- Medio Porte\n'
            '3- Grande Porte:\n'
            )
        
        
        #Calculando o valor da rota, de acordo com o modo de transporte escolhido.
        if caminhao == '1':
            valor = pp.calcular_rota(distanciaCidades)
            utilizado = pp
        elif caminhao == '2':
            valor = mp.calcular_rota(distanciaCidades)
            utilizado = mp
        elif caminhao == '3':
            valor = gp.calcular_rota(distanciaCidades)
            utilizado = gp
        else:
            raise Exception ('Número informado não corresponde a nenhum caminhão')
        print(f'de {cidadePartida} para {cidadeDestino}, utilizando um {utilizado}, a distância é de {distanciaCidades} e o custo de R$: {valor}')
        #Atribuindo o valor FALSE para caso o usuário informe outra opção na variável 'ação_usuário', a funcionalidade de consulta não ser exibida
        

    #Segunda funcionalidade, realizar o cadastro de um carregamento
def cadastrar():
        #Solicitando um número de identificação do transporte a ser realizado
        global iden
        #Criando uma chave com o número de identificação, para armazenar os dados
        dicionario_carregamentos[iden] = {}
        #Variável 'informar_cidades' criada como True para o usuário informar a quantidade de cidades de sua escolha.
        informar_cidades = True
        lista_cidades = []
        inicial = input('Informe a cidade de onde a frota sairá: ').upper()
        
        
        
        #Início do Loop para o usuário informar as cidades que ele deseja que o carregamento passe
        while informar_cidades:
            cidades = input(
                'Informe, em sequência, as cidades de destino desejadas, ao finalizar digite F: \n').upper()
            if cidades == 'F':
                informar_cidades = False
                break
            elif cidades not in excel[0]:
                print(f'Cidade {cidades} não está no sistema, informe corretamente ')
                continue
            
            lista_cidades.append(cidades)
        #Capturando a distância entre as cidades e juntando elas, para encontrar a distância total que o carregamento irá percorrer.
        cidade1 = inicial
        distancia = 0
        for i in lista_cidades:
            indexInicial = excel[0].index(cidade1)
            indexProxima = excel[0].index(i)
            distancia += excel[indexInicial + 1][indexProxima]
            cidade1 = i
            
            
        #Armazenando a distância total do carregamento
        
        dicionario_carregamentos[iden] = {
            'cidade_origem' : inicial,
            'cidade_destino' : lista_cidades[-1],
            'distancia' : distancia,
            'itens' : []
        }
        #Variável criada com valor TRUE pela mesma razão da variável 'informar_cidades'
        informar_itens = True
        
        itens_lista = []
        itens_quantidade = []
        itens_peso = []
        peso_unitario = 0
        peso_geral = 0
        #Criando uma lista dos itens, para depois serem atribuidos os nomes e suas respectivas quantidades
        
        while informar_itens:
            item = input(
                'Informe o nome do item que deseja adicionar ao carregamento(Caso queira finalizar, digite F): ').upper()

            if item == 'F':
                break
            itens_lista.append(item.upper())
            peso_item = input('Informe o peso da unidade deste item(em Kg): ')
            quantidade_item = input(
                'Informe a quantidade de unidades desse item: ')
            itens_quantidade.append(quantidade_item)
            #Variável peso_unitario armazena o peso total de um item específico, por exemplo, o peso de todas bananas combinadas
            peso_unitario = (float(peso_item) * int(quantidade_item))
            itens_peso.append(peso_unitario)
            #Variável peso_geral armazena o peso de todos os itens do carregamento
            peso_geral += peso_unitario
            #Criação de chave com nome do item e atribuindo o valor de quantidade presente no carregamento.
            dicionario_carregamentos[iden]['itens'].append(item)
            dicionario_carregamentos[iden]['itens'][-1] = { }
            dicionario_carregamentos[iden]['itens'][-1][item] = quantidade_item
            

        custo_parcial = 0
        distancia_parada = 0
        peso_restante = 0
        caminhao_parada = []
        #Validando se há mais de 1 cidade de destino (a cidade de origem não está na lista), para realizar a parada e o descarregamento.
        if len(lista_cidades) > 1:
            peso_carregamento = 0
            cidade_parada = input(
                'Informe em que cidade será feita a parada: ').upper()
            #Capturando o index da cidade que os caminhões vão parar, para calcular a distância restante de viagem após a parada
            idx_parada = lista_cidades.index(cidade_parada)
            #Armazenando os índices dos nomes das cidades e utilizando eles para capturar a distância entre elas, e somando toda distância a ser percorrida
            indexInicial = excel[0].index(inicial)
            indexProxima = excel[0].index(cidade_parada)
            distancia_parada += excel[indexInicial + 1][indexProxima]

            #Peso_carregamento é o peso a ser zerado, variável utilizada para atribuir o peso dos itens aos caminhões
            peso_carregamento = peso_geral

            #Alocando o peso nos caminhões e adicionando eles a lista de veículos utilizados
            while peso_carregamento > 0:

                if peso_carregamento >= gp.capacidade:
                    caminhao_parada.append(gp)
                    peso_carregamento = peso_carregamento - gp.capacidade

                elif peso_carregamento > mp.capacidade*2 and peso_carregamento <= gp.capacidade:
                    caminhao_parada.append(gp)
                    peso_carregamento = peso_carregamento - gp.capacidade

                elif peso_carregamento >= mp.capacidade:
                    caminhao_parada.append(mp)
                    peso_carregamento = peso_carregamento - mp.capacidade

                elif peso_carregamento > pp.capacidade*2 and peso_carregamento <= mp.capacidade:
                    caminhao_parada.append(mp)
                    peso_carregamento = peso_carregamento - mp.capacidade

                else:
                    caminhao_parada.append(pp)
                    peso_carregamento = peso_carregamento - pp.capacidade

            
            #Exibindo os itens que estão no carregamento e suas quantidades para o usuário selecionar se gostaria de descarregar ou não
            for i in itens_lista:
                
                descarregar = input(
                    f'Deseja descarregar {i} ?(S/N):  ').upper()

                if descarregar == 'N':
                    continue
                if descarregar == 'S':
                    idx_item = itens_lista.index(i)
                    qtd_descarregar = input(
                    f'Quantidade de {i} no carregamento: {itens_quantidade[idx_item]} Informe a quantidade desse item que deseja descarregar: ')
                    if qtd_descarregar > itens_quantidade[idx_item]:
                        print('Você informou um valor maior do que o disponível para descarregar')
                    else:
                        #Calculando o peso de uma unidade, e depois removendo o peso dos itens descarregados
                        itens_peso[idx_item] -= (float(itens_peso[idx_item]) / int(itens_quantidade[idx_item])) * int(qtd_descarregar)
                        #Removendo o número de itens que foram descarregados
                        itens_quantidade[idx_item] = int(itens_quantidade[idx_item]) - int(qtd_descarregar)
            
            #Somando o peso que sobrou para o restante da viagem
            for i in itens_peso:
                peso_restante += float(i)

            #Calculando o custo dos caminhões até esse trecho da viagem
            for i in caminhao_parada:
                custo_parcial += i.calcular_rota(distancia_parada)
                
            print(f'Percuso {inicial} até a cidade {cidade_parada}')
            print('Custo parcial até a parada de:', "%.2f" %
                  custo_parcial, '\nCaminhões utilizados: ')
            qtd_pp = caminhao_parada.count(pp)
            qtd_mp = caminhao_parada.count(mp)
            qtd_gp = caminhao_parada.count(gp)
            if qtd_pp > 0:
                print(qtd_pp, pp, '\n')
            if qtd_mp > 0:
                print(qtd_mp, mp, '\n')
            if qtd_gp > 0:
                print(qtd_gp, gp)
            
            #Atualizando a lista de cidades apenas com as cidades restantes no trajeto
            lista_cidades = lista_cidades[idx_parada:]
        
        #Distancia restante na viagem atualizada, caso não tenha a parada, a variável 'distancia_parada' tem valor 0, então não ira afetar a distância total
        distancia = distancia - distancia_parada
        caminhoes_carregamento = []
        
        #Atribuindo o peso restante aos caminhões e verificando se algum pode ser dispensado. Se não houver a parada, o código pula essa parte para realizar a mesma função
        #porém sem verificar se o caminhão ja estava sendo utilizando e sem a possibilidade de excluir nenhum caminhão
        if caminhao_parada:
            peso_total = peso_restante
            while peso_total > 0:

                if peso_total >= gp.capacidade and gp in caminhao_parada:

                    caminhoes_carregamento.append(gp)
                    caminhao_parada.remove(gp)
                    peso_total = peso_total - gp.capacidade

                elif peso_total > mp.capacidade*2 and peso_total <= gp.capacidade and gp in caminhao_parada:

                    caminhoes_carregamento.append(gp)
                    caminhao_parada.remove(gp)

                    peso_total = peso_total - gp.capacidade

                elif peso_total >= mp.capacidade and mp in caminhao_parada:

                    caminhoes_carregamento.append(mp)
                    caminhao_parada.remove(mp)

                    peso_total = peso_total - mp.capacidade

                elif peso_total > pp.capacidade*2 and peso_total <= mp.capacidade and mp in caminhao_parada:

                    caminhoes_carregamento.append(mp)
                    caminhao_parada.remove(mp)

                    peso_total = peso_total - mp.capacidade

                elif pp in caminhao_parada:

                    caminhoes_carregamento.append(pp)
                    caminhao_parada.remove(pp)

                    peso_total = peso_total - pp.capacidade
                else:
                    caminhoes_carregamento = caminhao_parada
                    peso_total = 0
        else:
            peso_total = peso_geral
            while peso_total > 0:

                if peso_total >= gp.capacidade:

                    caminhoes_carregamento.append(gp)

                    peso_total = peso_total - gp.capacidade

                elif peso_total > mp.capacidade*2 and peso_total <= gp.capacidade:

                    caminhoes_carregamento.append(gp)

                    peso_total = peso_total - gp.capacidade

                elif peso_total >= mp.capacidade:

                    caminhoes_carregamento.append(mp)

                    peso_total = peso_total - mp.capacidade

                elif peso_total > pp.capacidade*2 and peso_total <= mp.capacidade:

                    caminhoes_carregamento.append(mp)

                    peso_total = peso_total - mp.capacidade

                else:

                    caminhoes_carregamento.append(pp)

                    peso_total = peso_total - pp.capacidade
        
        #calculando os custos totais e salvando outras informações sobre a viagem no dicionário criado para armazenar esses dados
        custo = 0
        dicionario_carregamentos[iden]['caminhoes_utilizados'] = []
        for i in caminhoes_carregamento:
            custo += i.calcular_rota(distancia)
            dicionario_carregamentos[iden]['caminhoes_utilizados'].append(i)
        #Calculando o custo por km rodado
        dicionario_carregamentos[iden]['custo_por_km'] = "%.2f" % ((custo + custo_parcial) / (distancia + distancia_parada))
        qtd_pp, qtd_mp, qtd_gp = caminhoes_carregamento.count(pp), caminhoes_carregamento.count(mp),caminhoes_carregamento.count(gp)
         
        print(f'Percuso {inicial} até a cidade {lista_cidades[-1]}')
        print('Custo total de:', "%.2f" %
              (custo + custo_parcial), '\nCaminhões utilizados:\n')
        if qtd_pp > 0:
            print(qtd_pp, pp, '\n')
        if qtd_mp > 0:
            print(qtd_mp, mp, '\n')
        if qtd_gp > 0:
            print(qtd_gp, gp)
        dicionario_carregamentos[iden]['custo_total'] = "%.2f" % (
            custo + custo_parcial)
        #Armazenando a quantidade de veículos que foram utilizados nessa viagem
        dicionario_carregamentos[iden]['veiculos_utilizados'] = qtd_gp + qtd_mp + qtd_pp
        #Atribuindo o valor FALSE para não se repetir esse trecho do código caso o usuário não queira
        
        iden += 1
    #Exibindo os dados que foram salvos durante o cadastro das viagens
def relatorio():
        #Passando pelo dicionário para capturar os dados salvos anteriormente
        for i in dicionario_carregamentos:
            total_itens = 0
            print(f'ID do carregamento: {i}\nCidade origem:', dicionario_carregamentos[i]['cidade_origem'],'\nCidade Destino:', dicionario_carregamentos[i]['cidade_destino'], '\nDistância percorrida:\
',dicionario_carregamentos[i]['distancia'],'KM\nCusto total: R$',dicionario_carregamentos[i]['custo_total'], '\nVeículos utilizados: ', dicionario_carregamentos[i]['veiculos_utilizados'],'\
\nCusto por Km: R$', dicionario_carregamentos[i]['custo_por_km'])
            #Passando pela lista dentro da chave 'caminhoes_utlizados'
            qtd_peq, qtd_med, qtd_gra = dicionario_carregamentos[i]['caminhoes_utilizados'].count(pp), dicionario_carregamentos[i]['caminhoes_utilizados'].count(mp) ,dicionario_carregamentos[i]['caminhoes_utilizados'].count(gp)
            if qtd_peq > 0:
                print(f'{qtd_peq} - Caminhão de pequeno porte')
            if qtd_med > 0:
                print(f'{qtd_med} - Caminhão de médio porte')
            if qtd_gra > 0:
                print(f'{qtd_gra} - Caminhão de grande porte')
            print('------------------------------')
            #Passando pela lista dentro da chave 'itens'
            for p in dicionario_carregamentos[i]['itens']:
                #Capturando as keys e os values da lista(Nome do item e quantidade)
                for c, v in p.items():
                    print(f'Item: {c}, Quantidade: {v}') 
                    preco_por_item =  (float(dicionario_carregamentos[i]['custo_total']) / int(v))
                    print(f'Preço por {c}: R$', "%.2f" % preco_por_item)
                    total_itens += int(v)
            preco_por_unidade = float(dicionario_carregamentos[i]['custo_total']) / total_itens
            print(f'Total de itens transportados: {total_itens}\n'
                  f'Preço por unidade transportada: R$ {"%.2f" % preco_por_unidade}')
        
        #Atribuindo o valor FALSE para o trecho ser exibido somente quando o usuário solicitar
       

iden = 1
while True:
    acao_usuario = input(
        'Qual funcionalidade deseja utilizar?\n1 - Consulta\n2 - Cadastrar Transporte\n3 - Exibir relatorio\n4 - Finalizar programa: ')
    #Capturando o que o usuário deseja fazer
    if acao_usuario == '1':
        consulta()
    elif acao_usuario == '2':
        cadastrar()
    elif acao_usuario == '3':
        relatorio()
    elif acao_usuario == '4':
        break
    else:
        print('Ação informada é incorreta')
        continue